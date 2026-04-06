# ════════════════════════════════════════════════════════
# ÉTAPE 2 — Générer l'Excel de mapping images ↔ produits
# À exécuter APRÈS l'étape 1 (quand scan_images_vercel.json existe)
# ════════════════════════════════════════════════════════
#
# Installer d'abord : pip install openpyxl
#

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Charger le scan ───
json_path = r"C:\DATA-MC-2030\97IMPORT\scan_images_vercel.json"
with open(json_path, "r", encoding="utf-8") as f:
    scan = json.load(f)

images = scan["images"]

# ─── Regrouper les images par dossier parent ───
by_folder = {}
for img in images:
    folder = img["parent_folder"]
    if folder not in by_folder:
        by_folder[folder] = []
    by_folder[folder].append(img)

# ─── Créer le workbook ───
wb = Workbook()

# ═══════════════════════════════════════
# FEUILLE 1 : Toutes les images brutes
# ═══════════════════════════════════════
ws1 = wb.active
ws1.title = "Toutes les Images"

# En-têtes
headers = ["#", "Dossier Parent", "Nom Fichier", "Chemin Relatif", "Taille (Ko)", "Produit Associé (à remplir)"]
header_fill = PatternFill("solid", fgColor="1E3A5F")
header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

# Données
for i, img in enumerate(images, 1):
    row = i + 1
    ws1.cell(row=row, column=1, value=i).border = thin_border
    ws1.cell(row=row, column=2, value=img["parent_folder"]).border = thin_border
    ws1.cell(row=row, column=3, value=img["filename"]).border = thin_border
    ws1.cell(row=row, column=4, value=img["relative_path"]).border = thin_border
    ws1.cell(row=row, column=5, value=img["size_kb"]).border = thin_border
    # Colonne vide pour que Michel puisse remplir manuellement si besoin
    cell = ws1.cell(row=row, column=6, value="")
    cell.fill = PatternFill("solid", fgColor="FFFDE7")
    cell.border = thin_border

# Largeurs
ws1.column_dimensions["A"].width = 5
ws1.column_dimensions["B"].width = 18
ws1.column_dimensions["C"].width = 35
ws1.column_dimensions["D"].width = 55
ws1.column_dimensions["E"].width = 12
ws1.column_dimensions["F"].width = 30

# ═══════════════════════════════════════
# FEUILLE 2 : Regroupement par dossier
# ═══════════════════════════════════════
ws2 = wb.create_sheet("Par Dossier")

headers2 = ["Dossier", "Nb Images", "Liste Fichiers"]
for col, header in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

row = 2
for folder, folder_imgs in sorted(by_folder.items()):
    ws2.cell(row=row, column=1, value=folder).border = thin_border
    ws2.cell(row=row, column=2, value=len(folder_imgs)).border = thin_border
    filenames = "\n".join([img["filename"] for img in folder_imgs])
    cell = ws2.cell(row=row, column=3, value=filenames)
    cell.alignment = Alignment(wrap_text=True)
    cell.border = thin_border
    row += 1

ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 60

# ═══════════════════════════════════════
# FEUILLE 3 : Mapping Produit ↔ Images
# (pré-rempli par détection automatique)
# ═══════════════════════════════════════
ws3 = wb.create_sheet("Mapping Produits")

# Liste des produits connus (basée sur le catalogue Firestore)
# Michel pourra ajuster cette liste
produits_connus = [
    {"numero_interne": "MP-R10-001", "nom": "R10", "categorie": "mini-pelles", "keywords": ["r10", "R10"]},
    {"numero_interne": "MP-R13-001", "nom": "R13 PRO", "categorie": "mini-pelles", "keywords": ["r13", "R13"]},
    {"numero_interne": "MP-R15-001", "nom": "R15 PRO", "categorie": "mini-pelles", "keywords": ["r15", "R15"]},
    {"numero_interne": "MP-R18-001", "nom": "R18 PRO", "categorie": "mini-pelles", "keywords": ["r18", "R18"]},
    {"numero_interne": "MP-R20-001", "nom": "R20", "categorie": "mini-pelles", "keywords": ["r20", "R20"]},
    {"numero_interne": "MP-R22-001", "nom": "R22 PRO", "categorie": "mini-pelles", "keywords": ["r22", "R22"]},
    {"numero_interne": "MP-R25-001", "nom": "R25 PRO", "categorie": "mini-pelles", "keywords": ["r25", "R25"]},
    {"numero_interne": "MP-R30-001", "nom": "R30", "categorie": "mini-pelles", "keywords": ["r30", "R30"]},
    {"numero_interne": "MP-R32-001", "nom": "R32 PRO", "categorie": "mini-pelles", "keywords": ["r32", "R32"]},
    {"numero_interne": "MP-R35-001", "nom": "R35 PRO", "categorie": "mini-pelles", "keywords": ["r35", "R35"]},
    {"numero_interne": "KS-10K-001", "nom": "Kit Solaire 10 kW", "categorie": "solaire", "keywords": ["solar", "10kw", "10KW"]},
    {"numero_interne": "KS-12K-001", "nom": "Kit Solaire 12 kW", "categorie": "solaire", "keywords": ["solar", "12kw", "12KW"]},
    {"numero_interne": "KS-20K-001", "nom": "Kit Solaire 20 kW", "categorie": "solaire", "keywords": ["solar", "20kw", "20KW"]},
    {"numero_interne": "MP-20-001", "nom": "Maison Modulaire Premium", "categorie": "maisons", "keywords": ["maison", "modulaire", "premium", "house"]},
    {"numero_interne": "MS-20-001", "nom": "Maison Modulaire Standard", "categorie": "maisons", "keywords": ["maison", "modulaire", "standard", "house"]},
    {"numero_interne": "CC-BYD-001", "nom": "Camping Car Deluxe Hybride", "categorie": "maisons", "keywords": ["camping", "byd", "car"]},
    {"numero_interne": "ACC-AR-001", "nom": "Attache rapide Hydraulique", "categorie": "accessoires", "keywords": ["attache", "rapide", "quick"]},
    {"numero_interne": "ACC-GD-001", "nom": "Godet à dents", "categorie": "accessoires", "keywords": ["godet", "dents", "bucket"]},
    {"numero_interne": "ACC-GI-001", "nom": "Godet inclinable", "categorie": "accessoires", "keywords": ["godet", "inclinable", "tilt"]},
    {"numero_interne": "ACC-GC-001", "nom": "Godet de curage", "categorie": "accessoires", "keywords": ["godet", "curage", "cleaning"]},
    {"numero_interne": "ACC-GP-001", "nom": "Grappin", "categorie": "accessoires", "keywords": ["grappin", "grapple"]},
    {"numero_interne": "ACC-MH-001", "nom": "Marteau hydraulique", "categorie": "accessoires", "keywords": ["marteau", "hammer", "destructeur"]},
]

# En-têtes feuille 3
headers3 = [
    "N° Interne", "Nom Produit", "Catégorie",
    "Image 1 (chemin local)", "Image 2", "Image 3", "Image 4",
    "Nb Images Trouvées", "Statut",
    "Nouveau nom 1", "Nouveau nom 2", "Nouveau nom 3", "Nouveau nom 4"
]
for col, header in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = thin_border

# Matching automatique
ok_fill = PatternFill("solid", fgColor="C8E6C9")      # vert clair
warn_fill = PatternFill("solid", fgColor="FFF9C4")     # jaune clair
error_fill = PatternFill("solid", fgColor="FFCDD2")    # rouge clair

row = 2
for produit in produits_connus:
    ni = produit["numero_interne"]
    nom = produit["nom"]
    cat = produit["categorie"]
    keywords = produit["keywords"]

    # Chercher des images qui matchent par keyword dans le nom de fichier ou dossier parent
    matched_images = []
    for img in images:
        fname_lower = img["filename"].lower()
        folder_lower = img["parent_folder"].lower()
        path_lower = img["relative_path"].lower()
        
        for kw in keywords:
            kw_lower = kw.lower()
            if kw_lower in fname_lower or (kw_lower in folder_lower and cat.lower()[:4] in path_lower):
                if img not in matched_images:
                    matched_images.append(img)
                break

    nb_found = len(matched_images)
    
    # Écrire la ligne
    ws3.cell(row=row, column=1, value=ni).border = thin_border
    ws3.cell(row=row, column=2, value=nom).border = thin_border
    ws3.cell(row=row, column=3, value=cat).border = thin_border

    for i in range(4):
        cell_img = ws3.cell(row=row, column=4 + i)
        cell_new = ws3.cell(row=row, column=10 + i)
        if i < nb_found:
            cell_img.value = matched_images[i]["relative_path"]
            cell_new.value = f"{ni}-{str(i+1).zfill(2)}.png"
        else:
            cell_img.value = ""
            cell_new.value = f"{ni}-{str(i+1).zfill(2)}.png"
        cell_img.border = thin_border
        cell_new.border = thin_border

    ws3.cell(row=row, column=8, value=nb_found).border = thin_border
    
    status_cell = ws3.cell(row=row, column=9)
    if nb_found >= 2:
        status_cell.value = "✅ OK"
        status_cell.fill = ok_fill
    elif nb_found == 1:
        status_cell.value = "⚠️ 1 seule image"
        status_cell.fill = warn_fill
    else:
        status_cell.value = "❌ AUCUNE IMAGE"
        status_cell.fill = error_fill
    status_cell.border = thin_border

    row += 1

# Largeurs feuille 3
col_widths_3 = [14, 28, 14, 45, 45, 45, 45, 10, 18, 22, 22, 22, 22]
for i, w in enumerate(col_widths_3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ═══════════════════════════════════════
# FEUILLE 4 : Instructions
# ═══════════════════════════════════════
ws4 = wb.create_sheet("Instructions")
instructions = [
    ["═══ INSTRUCTIONS — REMISE EN ORDRE DES IMAGES 97IMPORT ═══"],
    [""],
    ["SITUATION :"],
    ["Les images existent en local dans le dossier vercel mais ne sont pas"],
    ["correctement liées aux produits dans Firebase après la migration."],
    [""],
    ["ÉTAPES À SUIVRE :"],
    [""],
    ["1. Vérifier l'onglet 'Mapping Produits' — les images sont-elles bien associées ?"],
    ["   - ✅ OK = au moins 2 images trouvées automatiquement"],
    ["   - ⚠️ = 1 seule image, vérifier manuellement"],
    ["   - ❌ = aucune image trouvée, remplir manuellement la colonne 'Image 1'"],
    [""],
    ["2. Pour les produits ❌, chercher dans l'onglet 'Toutes les Images'"],
    ["   et copier le chemin relatif dans la bonne ligne de 'Mapping Produits'"],
    [""],
    ["3. Une fois le mapping vérifié, donner ce fichier Excel à Claude Code"],
    ["   avec le prompt : 'Upload les images dans Firebase Storage selon le mapping'"],
    [""],
    ["CONVENTION DE NOMMAGE FIREBASE STORAGE :"],
    ["  products/{firestore_id}/{numero_interne}-01.png"],
    ["  products/{firestore_id}/{numero_interne}-02.png"],
    ["  etc."],
    [""],
    ["DOSSIER SOURCE LOCAL :"],
    [r"C:\DATA-MC-2030\97IMPORT\97import2026_siteweb\vercel"],
    [""],
    ["FIREBASE STORAGE BUCKET :"],
    ["import-412d0.firebasestorage.app"],
]

for r, row_data in enumerate(instructions, 1):
    cell = ws4.cell(row=r, column=1, value=row_data[0])
    if r == 1:
        cell.font = Font(bold=True, name="Arial", size=13, color="1E3A5F")
    elif row_data[0] in ["SITUATION :", "ÉTAPES À SUIVRE :", "CONVENTION DE NOMMAGE FIREBASE STORAGE :", "DOSSIER SOURCE LOCAL :", "FIREBASE STORAGE BUCKET :"]:
        cell.font = Font(bold=True, name="Arial", size=11, color="166534")

ws4.column_dimensions["A"].width = 80

# ─── Sauvegarder ───
output_path = r"C:\DATA-MC-2030\97IMPORT\CATALOGUE-MAPPING-IMAGES.xlsx"
wb.save(output_path)

print(f"\n{'═' * 60}")
print(f"✅ Excel généré : {output_path}")
print(f"{'═' * 60}")
print(f"📊 {len(images)} images scannées")
print(f"📦 {len(produits_connus)} produits référencés")

# Résumé matching
ok = sum(1 for p in produits_connus if len([img for img in images if any(kw.lower() in img['filename'].lower() for kw in p['keywords'])]) >= 2)
warn = sum(1 for p in produits_connus if len([img for img in images if any(kw.lower() in img['filename'].lower() for kw in p['keywords'])]) == 1)
err = len(produits_connus) - ok - warn
print(f"\n✅ {ok} produits avec images OK")
print(f"⚠️  {warn} produits avec 1 seule image")
print(f"❌ {err} produits sans image détectée")
print(f"\n👉 Ouvre le fichier Excel et vérifie l'onglet 'Mapping Produits'")
