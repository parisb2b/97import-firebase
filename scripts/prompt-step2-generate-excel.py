# -*- coding: utf-8 -*-
"""
ÉTAPE 2 — Génération CATALOGUE-MAPPING-IMAGES.xlsx
Lit : scripts/images-inventory.json + scripts/firestore-products.json
Génère : CATALOGUE-MAPPING-IMAGES.xlsx (à la racine du projet)
Lecture seule — ne modifie rien dans Firebase.

Onglets :
  1. Mapping Produits   — matching par mots-clés, couleurs vert/jaune/rouge
  2. Inventaire Images  — toutes les images trouvées par sous-dossier
  3. Images Sans Match  — images non associées à un produit
  4. Résumé Catégories  — synthèse par catégorie produit
"""

import json
import sys
import os
from pathlib import Path

# ── Vérifier openpyxl ──────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl manquant. Installation :", file=sys.stderr)
    print("   pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ── Chemins ────────────────────────────────────────────────────────────────
SCRIPT_DIR    = Path(__file__).parent
PROJECT_ROOT  = SCRIPT_DIR.parent
INVENTORY_JSON  = SCRIPT_DIR / "images-inventory.json"
PRODUCTS_JSON   = SCRIPT_DIR / "firestore-products.json"
OUTPUT_EXCEL    = PROJECT_ROOT / "CATALOGUE-MAPPING-IMAGES.xlsx"

# ── Couleurs ───────────────────────────────────────────────────────────────
VERT_REMPLI   = PatternFill("solid", fgColor="C6EFCE")  # ≥2 images
JAUNE_REMPLI  = PatternFill("solid", fgColor="FFEB9C")  # 1 image
ROUGE_REMPLI  = PatternFill("solid", fgColor="FFC7CE")  # 0 image
BLEU_ENTETE   = PatternFill("solid", fgColor="1E3A5F")  # navy admin
GRIS_ENTETE   = PatternFill("solid", fgColor="4A5568")
BLANC_FONT    = Font(color="FFFFFF", bold=True)
BOLD_FONT     = Font(bold=True)


# ── Règles de matching produit → images ───────────────────────────────────
# Pour chaque produit, on génère des mots-clés depuis son nom et numero_interne
# et on cherche dans l'inventaire les images dont le nom contient ces mots-clés.

CATEGORIE_KEYWORDS = {
    "miniexcavateur": ["r10", "r13", "r15", "r18", "r22", "r32", "r57"],
    "camping": ["camping", "camping_car"],
    "modulaire": ["modular", "modulaire"],
    "solaire": ["solar", "solaire", "kit"],
    "accessoire": ["godet", "fourche", "grappin", "marteau", "tariere", "ripper",
                   "rake", "auger", "flat_bucket", "tilt_bucket", "tooth_bucket",
                   "hydraulic", "grapple"],
}


def extraire_keywords_produit(produit: dict) -> list[str]:
    """Extrait les mots-clés de recherche depuis un produit Firestore."""
    keywords = set()

    # Depuis numero_interne (ex: "MP-R18-001" → "r18")
    num = produit.get("numero_interne", "").lower()
    for part in num.replace("-", " ").replace("_", " ").split():
        if len(part) >= 2:
            keywords.add(part)

    # Depuis le nom (ex: "Miniexcavatrice R18 Pro" → "r18", "pro")
    nom = produit.get("nom", "").lower()
    for part in nom.replace("-", " ").replace("_", " ").split():
        if len(part) >= 2:
            keywords.add(part)

    # Depuis la catégorie
    cat = produit.get("categorie", "").lower()
    for part in cat.replace("-", " ").replace("_", " ").split():
        if len(part) >= 2:
            keywords.add(part)

    # Mots génériques à ignorer
    STOP_WORDS = {"pro", "eco", "le", "la", "les", "de", "du", "des", "un", "une",
                  "et", "en", "au", "aux", "ma", "mp", "001", "002", "003"}
    keywords -= STOP_WORDS

    return sorted(keywords)


def matcher_images(produit: dict, images: list[dict]) -> list[dict]:
    """Retourne les images qui correspondent au produit (par mots-clés)."""
    keywords = extraire_keywords_produit(produit)
    if not keywords:
        return []

    matches = []
    for img in images:
        nom_lower = img["nom_sans_ext"].lower()
        chemin_lower = img["chemin_relatif"].lower()
        # Une image correspond si au moins un keyword apparaît dans son nom ou chemin
        if any(kw in chemin_lower for kw in keywords):
            matches.append(img)

    return matches


# ── Helpers Excel ──────────────────────────────────────────────────────────
def entete(ws, row: int, values: list, fill=BLEU_ENTETE):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.font = BLANC_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_width(ws, min_w=8, max_w=60):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)


# ── Onglet 1 : Mapping Produits ────────────────────────────────────────────
def onglet_mapping(wb, produits: list[dict], images: list[dict]):
    ws = wb.create_sheet("Mapping Produits")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    headers = [
        "#", "ID Firestore", "N° Interne", "Nom FR", "Catégorie",
        "Nb Images Trouvées", "Statut", "Mots-clés utilisés",
        "Image 1", "Image 2", "Image 3", "Image 4", "Image 5",
        "Sous-dossier principal",
    ]
    entete(ws, 1, headers)

    # Stats pour résumé
    stats = {"vert": 0, "jaune": 0, "rouge": 0}

    for idx, produit in enumerate(produits, 1):
        matches = matcher_images(produit, images)
        nb = len(matches)
        keywords = extraire_keywords_produit(produit)

        if nb >= 2:
            statut = "✅ OK"
            fill = VERT_REMPLI
            stats["vert"] += 1
        elif nb == 1:
            statut = "⚠️ 1 SEULE IMAGE"
            fill = JAUNE_REMPLI
            stats["jaune"] += 1
        else:
            statut = "❌ AUCUNE IMAGE"
            fill = ROUGE_REMPLI
            stats["rouge"] += 1

        # Sous-dossier dominant
        sous_dossiers = list(dict.fromkeys(m["sous_dossier"] for m in matches))
        sous_dossier_str = " / ".join(sous_dossiers[:3]) if sous_dossiers else ""

        row_data = [
            idx,
            produit.get("id", ""),
            produit.get("numero_interne", ""),
            produit.get("nom", ""),
            produit.get("categorie", ""),
            nb,
            statut,
            ", ".join(keywords[:6]),
        ] + [m["chemin_relatif"] for m in matches[:5]] + [""] * max(0, 5 - len(matches)) + [sous_dossier_str]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=idx + 1, column=col, value=val)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=False, vertical="center")

    auto_width(ws)
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["I"].width = 50
    ws.column_dimensions["J"].width = 50
    ws.column_dimensions["K"].width = 50

    return stats


# ── Onglet 2 : Inventaire Images ──────────────────────────────────────────
def onglet_inventaire(wb, images: list[dict]):
    ws = wb.create_sheet("Inventaire Images")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    headers = ["#", "Sous-dossier", "Nom fichier", "Extension", "Taille Ko", "Chemin relatif", "Mots-clés"]
    entete(ws, 1, headers, fill=GRIS_ENTETE)

    for idx, img in enumerate(images, 1):
        row = [
            idx,
            img["sous_dossier"],
            img["nom"],
            img["extension"],
            img["taille_ko"],
            img["chemin_relatif"],
            ", ".join(img["mots_cles"]),
        ]
        for col, val in enumerate(row, 1):
            ws.cell(row=idx + 1, column=col, value=val).alignment = Alignment(vertical="center")

    auto_width(ws)
    ws.column_dimensions["F"].width = 55


# ── Onglet 3 : Images Sans Match ──────────────────────────────────────────
def onglet_sans_match(wb, images: list[dict], produits: list[dict]):
    # Trouver les images qui ne matchent aucun produit
    images_matchees = set()
    for produit in produits:
        for img in matcher_images(produit, images):
            images_matchees.add(img["chemin_relatif"])

    sans_match = [img for img in images if img["chemin_relatif"] not in images_matchees]

    ws = wb.create_sheet("Images Sans Match")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    headers = ["#", "Sous-dossier", "Nom fichier", "Extension", "Taille Ko", "Chemin relatif"]
    entete(ws, 1, headers, fill=PatternFill("solid", fgColor="7C3AED"))  # violet

    for idx, img in enumerate(sans_match, 1):
        row = [idx, img["sous_dossier"], img["nom"], img["extension"], img["taille_ko"], img["chemin_relatif"]]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=idx + 1, column=col, value=val)
            cell.alignment = Alignment(vertical="center")
            if idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F3E8FF")

    auto_width(ws)
    ws.column_dimensions["F"].width = 55

    # Note en haut
    ws.insert_rows(1)
    ws.cell(row=1, column=1, value=f"⚠️ {len(sans_match)} images non associées à un produit Firestore").font = Font(bold=True, color="7C3AED")


# ── Onglet 4 : Résumé Catégories ──────────────────────────────────────────
def onglet_resume(wb, produits: list[dict], images: list[dict], stats_mapping: dict):
    ws = wb.create_sheet("Résumé Catégories")
    ws.row_dimensions[1].height = 30

    # Résumé global en haut
    ws.cell(row=1, column=1, value="RÉSUMÉ GLOBAL").font = Font(bold=True, size=13)
    ws.cell(row=2, column=1, value=f"Total produits Firestore :").font = BOLD_FONT
    ws.cell(row=2, column=2, value=len(produits))
    ws.cell(row=3, column=1, value="✅ Avec ≥2 images :").font = BOLD_FONT
    ws.cell(row=3, column=2, value=stats_mapping["vert"]).fill = VERT_REMPLI
    ws.cell(row=4, column=1, value="⚠️ Avec 1 seule image :").font = BOLD_FONT
    ws.cell(row=4, column=2, value=stats_mapping["jaune"]).fill = JAUNE_REMPLI
    ws.cell(row=5, column=1, value="❌ Sans image :").font = BOLD_FONT
    ws.cell(row=5, column=2, value=stats_mapping["rouge"]).fill = ROUGE_REMPLI
    ws.cell(row=6, column=1, value=f"Total images locales :").font = BOLD_FONT
    ws.cell(row=6, column=2, value=len(images))

    # Résumé par catégorie
    ws.cell(row=8, column=1, value="PAR CATÉGORIE").font = Font(bold=True, size=12)
    entete(ws, 9, ["Catégorie", "Nb Produits", "✅ OK", "⚠️ 1 image", "❌ Aucune", "% Complet"], fill=BLEU_ENTETE)

    categories: dict[str, dict] = {}
    for produit in produits:
        cat = produit.get("categorie") or "Sans catégorie"
        if cat not in categories:
            categories[cat] = {"total": 0, "vert": 0, "jaune": 0, "rouge": 0}
        categories[cat]["total"] += 1
        matches = matcher_images(produit, images)
        nb = len(matches)
        if nb >= 2:
            categories[cat]["vert"] += 1
        elif nb == 1:
            categories[cat]["jaune"] += 1
        else:
            categories[cat]["rouge"] += 1

    for row_idx, (cat, data) in enumerate(sorted(categories.items()), 10):
        total = data["total"]
        pct = round((data["vert"] / total) * 100) if total > 0 else 0
        row = [cat, total, data["vert"], data["jaune"], data["rouge"], f"{pct}%"]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if pct >= 80:
                cell.fill = VERT_REMPLI
            elif pct >= 40:
                cell.fill = JAUNE_REMPLI
            else:
                cell.fill = ROUGE_REMPLI

    auto_width(ws)
    ws.column_dimensions["A"].width = 30


# ── Main ───────────────────────────────────────────────────────────────────
def main():
    # Vérifier fichiers source
    if not INVENTORY_JSON.exists():
        print(f"❌ Inventaire introuvable : {INVENTORY_JSON}", file=sys.stderr)
        print("   → Lancer d'abord : prompt-step1-scan-images.py", file=sys.stderr)
        sys.exit(1)

    if not PRODUCTS_JSON.exists():
        print(f"❌ Produits Firestore introuvables : {PRODUCTS_JSON}", file=sys.stderr)
        print("   → Lancer d'abord : npx tsx scripts/export-products-json.ts", file=sys.stderr)
        sys.exit(1)

    # Charger données
    with open(INVENTORY_JSON, encoding="utf-8") as f:
        inventory_data = json.load(f)
    images = inventory_data["images"]

    with open(PRODUCTS_JSON, encoding="utf-8") as f:
        produits = json.load(f)

    print(f"[INFO] {len(images)} images chargees depuis l'inventaire")
    print(f"[INFO] {len(produits)} produits Firestore charges")

    # Créer le workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Supprimer la feuille vide par défaut

    stats_mapping = onglet_mapping(wb, produits, images)
    onglet_inventaire(wb, images)
    onglet_sans_match(wb, images, produits)
    onglet_resume(wb, produits, images, stats_mapping)

    wb.save(OUTPUT_EXCEL)

    print(f"\n[OK] Excel genere : {OUTPUT_EXCEL}")
    print(f"\nResultat matching :")
    print(f"  [OK]      Produits (>=2 images) : {stats_mapping['vert']}")
    print(f"  [WARNING] 1 seule image          : {stats_mapping['jaune']}")
    print(f"  [ERROR]   Aucune image           : {stats_mapping['rouge']}")


if __name__ == "__main__":
    main()
